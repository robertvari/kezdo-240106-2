from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

ws['A1'] = "Robert"
ws["B1"] = "robert@gmail.com"
ws["C1"] = "Budapest"

ws['A2'] = "Csaba"
ws["B2"] = "csaba@gmail.com"
ws["C2"] = "Pécs"

# Save the file
wb.save("sample.xlsx")