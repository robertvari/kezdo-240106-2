import os, json
from my_functions.file_functions import get_all_files
from PIL import Image
from PIL.ExifTags import TAGS
from openpyxl import Workbook


def main():
    clear_screen()
    print("-"*50, "Get Proto Data", "-"*50)

    folder_path = get_root_folder()
    image_list = collect_photos(folder_path)
    photo_data = collect_data(image_list)
    
    save_json(photo_data)
    save_excel(photo_data)

def get_root_folder():
    # folder_path = input("Where are your photos?")
    folder_path = r"D:\Work\_PythonSuli\kezdo-240106\alapok_1\photos"

    assert os.path.isdir(folder_path), f"This is not a valid folder: {folder_path}"

    return folder_path

def collect_photos(folder_path):
    file_list = []
    get_all_files(folder_path, file_list, extension=".jpg")
    return file_list

def collect_data(image_list):
    photo_data = {}

    for photo in image_list:
        img = Image.open(photo)
        photo_data[photo] = {"width": img.size[0], "height": img.size[1], "date": None, "camera": None, "ISO": None, "GPSInfo": None}

        exif_data = img._getexif()
        if not exif_data:
            continue

        for tag in exif_data:
            if TAGS[tag] == "Model":
                photo_data[photo]["camera"] = exif_data[tag]
            
            if TAGS[tag] == "ISOSpeedRatings":
                photo_data[photo]["ISO"] = exif_data[tag]

            if TAGS[tag] == "DateTimeOriginal":
                photo_data[photo]["date"] = exif_data[tag]

            if TAGS[tag] == "GPSInfo":
                photo_data[photo]["GPSInfo"] = exif_data[tag]
                
    return photo_data

def save_json(photo_data):
    with open("photo_data.json", "w") as file:
        json.dump(photo_data, file)
    
    print("Photo data saved.")

def save_excel(photo_data):
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "Path"
    ws["B1"] = "Width"
    ws["C1"] = "Height"
    ws["D1"] = "Date"
    ws["E1"] = "Model"
    ws["F1"] = "ISO"

    counter = 3
    for photo in photo_data:
        ws[f"A{counter}"] = photo
        ws[f"B{counter}"] = photo_data[photo]["width"]
        ws[f"C{counter}"] = photo_data[photo]["height"]
        ws[f"D{counter}"] = photo_data[photo]["date"]
        ws[f"E{counter}"] = photo_data[photo]["camera"]
        ws[f"F{counter}"] = photo_data[photo]["ISO"]
        counter += 1

    wb.save("photo_data.xlsx")

def clear_screen():
    os.system("cls")  #linux, mac: os.system("clear")


main()